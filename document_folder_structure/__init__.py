"""
Markdown Documentation and Directory Tree Generator for Projects.

Usage example:
    python document_folder_structure.py C:\\path\\to\\your\\project

Creates two files in ./INTERNAL_DOCS/: 
    - directory_tree.md (directory structure with Git info)
    - all_files_content.md (index and file contents, binaries omitted)

Includes Git metadata (branch, commit, branches, repo) and GitHub links (if applicable).
Skips configurable folders/files (README, LICENSE, CHANGELOG, CONTRIBUTING, SECURITY).
Automatically adds created folder to .gitignore.

Sample config file, first 4 lines are for exclude files and directories before the tree generation, 5-11 lines are for truncate file content on Markdown documentation:
{
    "excluded_dirs": ["node_modules", "__pycache__", ".git", "deps", ".fingerprint", "build", "incremental"],
    "excluded_files": [".gitignore", ".849C9593-D756-4E56-8D6E-42412F2A707B"],
    "excluded_extensions": [".log", ".tmp", ".bak", ".db-shm", ".db-wal", ".npmrc", ".prettierignore", ".prettierrc"],
    "max_depth": 4,
    "truncate_files": ["README", "LICENSE", "CHANGELOG", "CONTRIBUTING", "SECURITY"],
    "truncate_exts": [".md", ".txt"],
    "truncate_dirs": ["docs"],
    "truncate_file_pairs": [["document_folder_structure", ".html"]],
    "max_truncate_chars": 2000,
    "truncate_lines": 10,
    "max_log_lines": 10,
    "max_preview_columns": 20
}

Requirements:
    - Python 3.6+
    - Run "pip install -r requirements.txt" to install dependencies.

"""
import csv
try:
    import openpyxl
except ImportError:
    openpyxl = None
import os
import json
import logging
import datetime
import subprocess
from argparse import ArgumentParser

DEFAULTS = {
    "excluded_dirs": ["node_modules", "__pycache__", ".git", "deps", ".fingerprint", "build", "incremental"],
    "excluded_files": [".gitignore", ".849C9593-D756-4E56-8D6E-42412F2A707B"],
    "excluded_extensions": [".log", ".tmp", ".bak", ".db-shm", ".db-wal", ".npmrc", ".prettierignore", ".prettierrc"],
    "max_depth": 4,
    "truncate_files": ["README", "LICENSE", "CHANGELOG", "CONTRIBUTING", "SECURITY"],
    "truncate_exts": [".md", ".txt"],
    "truncate_dirs": ["docs"],
    "truncate_file_pairs": [["document_folder_structure", ".html"]],
    "max_truncate_chars": 2000,
    "truncate_lines": 10,
    "max_log_lines": 10,
    "max_preview_columns": 20
}


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def add_line_numbers(lines, start=1):
    """
    Adds line numbers to each line in a list.
    """
    return [f"{i + start} | {line.rstrip()}\n" for i, line in enumerate(lines)]


def load_config_with_defaults(config_file, args):
    """
    Load config file.
    """
    config = DEFAULTS.copy()
    config_path = None
    
    if config_file and os.path.exists(config_file):
        config_path = config_file
    else:
        repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        root_config = os.path.join(repo_root, "config.json")
        if os.path.exists(root_config):
            config_path = root_config
    if config_path and os.path.exists(config_path):
        with open(config_path, 'r') as f:
            file_config = json.load(f)
        config.update(file_config)
    if getattr(args, "max_depth", None) is not None:
        config["max_depth"] = args.max_depth
    if getattr(args, "truncate_lines", None) is not None:
        config["truncate_lines"] = args.truncate_lines
    if getattr(args, "truncate_chars", None) is not None:
        config["max_truncate_chars"] = args.truncate_chars
    if getattr(args, "max_log_lines", None) is not None:
        config["max_log_lines"] = args.max_log_lines
    if getattr(args, "max_preview_columns", None) is not None:
        config["max_preview_columns"] = args.max_preview_columns
    return config


def read_n_lines_max_chars(filepath, config, max_lines=None, max_chars=None):
    """
    Reads up to `max_lines` lines and never more than `max_chars` characters.
    Returns a list of lines (not joined), possibly truncated.
    """
    
    if max_lines is None:
        max_lines = config["truncate_lines"]
    if max_chars is None:
        max_chars = config["max_truncate_chars"]
    
    lines = []
    char_count = 0
    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as fin:
            for i, line in enumerate(fin):
                if i >= max_lines or char_count >= max_chars:
                    break
                remaining = max_chars - char_count
                if len(line) > remaining:
                    lines.append(line[:remaining] + " ... [TRUNCATED]\n")
                    char_count += remaining
                    break
                else:
                    lines.append(line)
                    char_count += len(line)
    except Exception as e:
        lines = [f"(Could not read file: {e})\n"]
    return lines



def preview_csv(filepath, config, max_lines=None):
    """
    Returns the first lines of a CSV as a Markdown table.
    """
    
    if max_lines is None:
        max_lines = config["max_log_lines"]
    max_columns = config["max_preview_columns"]
    lines = []
    try:
        with open(filepath, newline='', encoding='utf-8', errors='replace') as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader, None)
            if headers:
                if len(headers) > max_columns:
                    lines.append(f"_(Only showing first {max_columns} columns)_")
                headers = headers[:max_columns]
                lines.append("| " + " | ".join(headers) + " |")
                lines.append("|" + "|".join("---" for _ in headers) + "|")
            count = 0
            for row in reader:
                row = row[:max_columns]
                lines.append("| " + " | ".join(row) + " |")
                count += 1
                if count >= max_lines:
                    break
    except Exception as e:
        lines = [f"(Could not read CSV: {e})"]
    return lines

def preview_excel(filepath, config, max_lines=None):
    """
    Returns the first lines of each sheet in an Excel (.xlsx) file as a Markdown table.
    """
    max_columns = config["max_preview_columns"]
    if max_lines is None:
        max_lines = config["max_log_lines"]
    
    
    if openpyxl is None:
        return ["(openpyxl is not installed, cannot preview Excel)"]
    lines = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            lines.append(f"### Sheet: {sheet}")
            headers = [str(h) if h is not None else "" for h in rows[0][:max_columns]]
            if len(rows[0]) > max_columns:
                lines.append(f"_(Only showing first {max_columns} columns)_")
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("|" + "|".join("---" for _ in headers) + "|")
            for i, row in enumerate(rows[1:max_lines+1]):
                cells = [str(cell) if cell is not None else "" for cell in row[:max_columns]]
                lines.append("| " + " | ".join(cells) + " |")
            lines.append("")
    except Exception as e:
        lines = [f"(Could not read Excel: {e})"]
    return lines


def ensure_gitignore_has_internal_docs(project_dir):
    """
    Ensures that the INTERNAL_DOCS/ folder is included in the .gitignore file.

    If .gitignore does not exist, it is created. If it exists but the line is missing, it is added.

    Args:
        project_dir (str): Path to the project root directory.
    """
    gitignore_path = os.path.join(project_dir, '.gitignore')
    try:
        if not os.path.exists(gitignore_path):
            with open(gitignore_path, 'w', encoding='utf-8') as f:
                f.write('INTERNAL_DOCS/\n')
            return
        with open(gitignore_path, 'r', encoding='utf-8') as f:
            lines = f.read().splitlines()
        if 'INTERNAL_DOCS/' not in lines:
            with open(gitignore_path, 'a', encoding='utf-8') as f:
                with open(gitignore_path, 'rb') as fr:
                    fr.seek(-1, os.SEEK_END)
                    last_char = fr.read(1)
                if last_char != b'\n':
                    f.write('\n')
                f.write('INTERNAL_DOCS/\n')
    except Exception as e:
        print(f"Warning: could not update .gitignore: {e}")



def ensure_internal_docs_dir(base_path):
    """
    Creates the INTERNAL_DOCS folder if it does not exist.

    Args:
        base_path (str): Base path to create the folder in.

    Returns:
        str: Absolute path to the INTERNAL_DOCS folder.
    """
    internal_docs = os.path.join(base_path, "INTERNAL_DOCS")
    if not os.path.exists(internal_docs):
        os.makedirs(internal_docs)
    return internal_docs

def count_tree_stats(directory, config, max_depth=None, depth=0):
    """
    Counts the total files and folders in the tree, respecting exclusions and depth.

    Args:
        directory (str): Path to the root directory.
        config (dict): Configuration (exclusions).
        max_depth (int, optional): Maximum depth to traverse.
        depth (int, optional): Current depth.

    Returns:
        (int, int): Tuple (files, folders)
    """
    total_files, total_dirs = 0, 0
    try:
        files = os.listdir(directory)
    except Exception:
        return 0, 0
    for file in files:
        path = os.path.join(directory, file)
        if os.path.isdir(path):
            if file not in config['excluded_dirs']:
                total_dirs += 1
                if max_depth is None or depth < max_depth:
                    f, d = count_tree_stats(path, config, max_depth, depth + 1)
                    total_files += f
                    total_dirs += d
        else:
            if file in config['excluded_files'] or file.endswith(tuple(config['excluded_extensions'])):
                continue
            total_files += 1
    return total_files, total_dirs

def is_binary(filename, blocksize=512):
    """
    Heuristically detects if a file is binary.

    Args:
        filename (str): File path.

    Returns:
        bool: True if the file appears to be binary.
    """
    try:
        with open(filename, 'rb') as f:
            chunk = f.read(blocksize)
            if b'\0' in chunk:
                return True
            try:
                chunk.decode('utf-8')
                return False
            except UnicodeDecodeError:
                return True
    except Exception:
        return True
    return False

def get_github_url(git_info, filepath):
    """
    Builds a GitHub (or GitLab) URL for a given file and commit.

    Args:
        git_info (tuple): (branch, commit, status, remote_url)
        filepath (str): Relative file path.

    Returns:
        str: GitHub URL (or empty if not applicable).
    """
    if not git_info or not git_info[3] or not git_info[1]:
        return ""
    remote = git_info[3]
    commit = git_info[1]
    # Supports SSH and HTTPS formats (GitHub and GitLab)
    if remote.startswith("git@"):
        remote = remote.replace(":", "/").replace("git@", "https://").replace(".git", "")
    elif remote.startswith("https://"):
        remote = remote.replace(".git", "")
    return f"{remote}/blob/{commit}/{filepath.replace(os.sep, '/')}"

def get_git_info(directory):
    """
    Gets Git information for the repository in the given directory.

    Args:
        directory (str): Path to the repo root directory.

    Returns:
        tuple: (branch, commit, status, remote_url), or (None, None, None, None) if not a repo.
    """
    try:
        branch = subprocess.check_output(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=directory, stderr=subprocess.DEVNULL
        ).decode().strip()
        commit = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=directory, stderr=subprocess.DEVNULL
        ).decode().strip()
        status = subprocess.check_output(
            ["git", "status", "--porcelain"],
            cwd=directory, stderr=subprocess.DEVNULL
        ).decode()
        dirty = "✔️ synced" if not status else "⚠️ local changes (need to push to branch)"
        remote = subprocess.check_output(
            ["git", "config", "--get", "remote.origin.url"],
            cwd=directory, stderr=subprocess.DEVNULL
        ).decode().strip()
        return branch, commit, dirty, remote
    except Exception:
        return None, None, None, None

def get_git_branches(directory):
    """
    Returns a list of existing branches in the repo.

    Args:
        directory (str): Repo path.

    Returns:
        list: List of branch names, empty if not a repo.
    """
    try:
        out = subprocess.check_output(
            ["git", "branch", "--format", "%(refname:short)"], cwd=directory
        ).decode().strip().split("\n")
        out = [x.strip() for x in out if x.strip()]
        return out
    except Exception:
        return []

# Extension dictionary for code highlighting in Markdown code blocks
LANG_EXT = {
    ".py": "python", ".md": "markdown", ".json": "json", ".js": "javascript",
    ".ts": "typescript", ".tsx": "tsx", ".jsx": "jsx", ".html": "html", ".css": "css",
    ".scss": "scss", ".sass": "sass", ".less": "less", ".yml": "yaml", ".yaml": "yaml",
    ".toml": "toml", ".sh": "bash", ".bat": "bat", ".ps1": "powershell", ".txt": "",
    ".xml": "xml", ".csv": "csv", ".ini": "ini", ".conf": "conf", ".php": "php", ".rb": "ruby",
    ".go": "go", ".java": "java", ".c": "c", ".cpp": "cpp", ".h": "cpp", ".hpp": "cpp", ".rs": "rust",
    ".env": "", ".sql": "sql", ".log": ""
}
"""
    Dictionary of possible file extensions for Markdown code highlighting.
"""
 
# Always exclude generated documentation files
DOC_OUTPUTS = {"directory_tree.md", "all_files_content.md"}

def generate_tree_and_collect_files(
    directory, prefix="", depth=0, max_depth=None, is_last=True, config=None,
    files_list=None, extensions_set=None, exclude_outputs=True, base_docs_path=None
):
    """
    Generates the directory tree (as text) and collects the list of valid files.

    Args:
        directory (str): Project root.
        prefix (str): Visual prefix for branches.
        depth (int): Current depth.
        max_depth (int): Maximum depth to traverse.
        is_last (bool): Whether this is the last item at this level.
        config (dict): Exclusion configuration.
        files_list (list): List to add found files to.
        extensions_set (set): Set of detected extensions.
        exclude_outputs (bool): Whether to exclude generated docs.
        base_docs_path (str): Base path for relative paths.

    Returns:
        list: Lines of the tree.
    """
    if max_depth is not None and depth > max_depth:
        return []
    try:
        files = os.listdir(directory)
    except PermissionError as e:
        logging.warning(f"Permission denied: {e}")
        return []
    files.sort()
    tree = []
    excluded_dirs = config['excluded_dirs']
    excluded_files = set(config['excluded_files'])
    excluded_extensions = tuple(config['excluded_extensions'])
    for index, file in enumerate(files):
        path = os.path.join(directory, file)
        is_last_item = index == len(files) - 1
        connector = "└── " if is_last_item else "├── "
        if os.path.isdir(path):
            if file in excluded_dirs:
                tree.append(f"{prefix}{connector}+ {file}/ (excluded)")
            else:
                tree.append(f"{prefix}{connector}+ {file}/")
                tree.extend(
                    generate_tree_and_collect_files(
                        path, prefix + ("    " if is_last_item else "│   "),
                        depth + 1, max_depth, is_last_item, config, files_list, extensions_set, exclude_outputs, base_docs_path
                    )
                )
        else:
            rel_file = os.path.relpath(path, base_docs_path)
            if (
                file in excluded_files or
                file.endswith(excluded_extensions) or
                (exclude_outputs and file in DOC_OUTPUTS and os.path.dirname(path).endswith("INTERNAL_DOCS"))
            ):
                continue
            tree.append(f"{prefix}{connector}- {file}")
            if files_list is not None:
                files_list.append(os.path.relpath(path, base_docs_path))
            if extensions_set is not None:
                ext = os.path.splitext(file)[1].lower()
                extensions_set.add(ext)
    return tree

def markdown_tree_header(title, total_files, total_dirs, git_info=None, extensions=None, branches=None):
    """
    Builds the header for the directory tree, with git metadata and extensions.

    Args:
        title (str): Title.
        total_files (int): Number of files.
        total_dirs (int): Number of folders.
        git_info (tuple): Git info.
        extensions (set): Detected extensions.
        branches (list): Repo branches.

    Returns:
        str: Markdown header.
    """
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header = f"# {title}\n\n_Generated: {now}_\n\n"
    if branches:
        header += f"**Existing branches:** {', '.join(branches)}  \n"  
    if git_info and git_info[0]:
        header += (
            f"**Current branch:** `{git_info[0]}`  \n"
            f"**Commit:** `{git_info[1]}`  \n"
            f"**Repo:** {git_info[3]}  \n"
            f"**Status:** {git_info[2]}\n\n"
        )
    if extensions:
        exts = ", ".join(sorted(ext for ext in extensions if ext))
        header += f"**Detected extensions:** `{exts}`\n\n" if exts else ""
    header += f"_(Files: {total_files}, Folders: {total_dirs})_\n\n```text\n"
    return header

def markdown_tree_footer():
    """Returns the footer for the tree block (code block closing)."""
    return "```\n"

def print_tree_and_collect_files(
    directory, config, output_file="directory_tree.md", tree_title="Project Directory Tree"
):
    """
    Prints the tree in Markdown and returns the list of found files and git_info.

    Returns:
        files_list (list): Found files.
        git_info (tuple): Git info of the repo.
    """
    max_depth = config.get('max_depth')
    files_list = []
    extensions_set = set()
    dir_name = os.path.basename(os.path.abspath(directory))    
    tree = generate_tree_and_collect_files(
        directory, max_depth=max_depth, config=config,
        files_list=files_list, extensions_set=extensions_set,
        exclude_outputs=True, base_docs_path=directory
    )
    total_files, total_dirs = count_tree_stats(directory, config, max_depth)
    git_info = get_git_info(directory)
    git_branches = get_git_branches(directory)
    header = markdown_tree_header(tree_title, total_files, total_dirs, git_info, extensions_set, git_branches)
    footer = markdown_tree_footer()
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(header)
        f.write(f"{dir_name}/\n")
        for line in tree:
            f.write(line + "\n")
        f.write(footer)
    return files_list, git_info

def group_files_by_folder(files_list):
    """
    Converts a list of files into a hierarchical tree (nested dicts) by folder.
    """
    tree = {}
    for path in files_list:
        parts = path.split(os.sep)
        cur = tree
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                cur.setdefault("_files", []).append(path)
            else:
                cur = cur.setdefault(part, {})
    return tree

def render_index(tree, basepath="", git_info=None, level=0):
    """
    Renders the grouped index in Markdown from the folder/file tree.

    Args:
        tree (dict): Tree generated by group_files_by_folder.
        basepath (str): Current base path.
        git_info (tuple): Git info for links.
        level (int): Indentation.

    Returns:
        list: Markdown lines.
    """
    lines = []
    indent = "    " * level
    folders = [k for k in tree if k != "_files"]
    for folder in folders:
        lines.append(f"{indent}- **{folder}/**")
        lines.extend(render_index(tree[folder], os.path.join(basepath, folder), git_info, level+1))
    if "_files" in tree:
        for filepath in tree["_files"]:
            anchor = filepath.replace("/", "").replace(".", "").replace("\\", "")
            url = get_github_url(git_info, filepath)
            fname = os.path.basename(filepath)
            link_line = f"{indent}    - [{fname}](#{anchor})"
            if url:
                link_line += f" | [GitHub]({url})"
            lines.append(link_line)
    return lines


    
    
def should_truncate_file(filepath, config):
    """
    Determines if a file should be truncated in preview, by name/extension/path.
    """
    
    for d in config["truncate_dirs"]:
        parts = [p.lower() for p in filepath.replace("\\", "/").split("/")]
        if d.lower() in parts:
            return True

    base = os.path.splitext(os.path.basename(filepath))[0]
    ext = os.path.splitext(filepath)[1].lower()
    for pair in config["truncate_file_pairs"]:
        if base == pair[0] and ext == pair[1]:
            return True

    if base.upper() in (t.upper() for t in config["truncate_files"]):
        return True

    if ext in config["truncate_exts"]:
        return True

    return False
    
    

def flatten_files_in_tree(tree):
    """
    Returns an ordered list of all files in the hierarchical tree.
    """
    files = []
    folders = [k for k in tree if k != "_files"]
    for folder in folders:
        files.extend(flatten_files_in_tree(tree[folder]))
    if "_files" in tree:
        files.extend(tree["_files"])
    return files

def generate_content_document(
    files_list, output_file, git_info, config, base_docs_path=None, project_dir=None, no_lines=False
):
    """
    Generates the content document with index and preview/truncation of each file.

    Args:
        files_list (list): List of files.
        output_file (str): Output file.
        git_info (tuple): Git info.
        config: Config.json file.
        base_docs_path (str): Root for paths.
        project_dir (str): For git branch info.

    Effect: writes the Markdown content file.
    """
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out_lines = [f"# Project File Contents\n\n_Generated: {now}_\n\n"]
    branches = get_git_branches(project_dir)
    if branches:
        out_lines.append(f"**Existing branches:** {', '.join(branches)}  \n")
    if git_info and git_info[0]:
        out_lines.append(
            f"**Current branch:** `{git_info[0]}`  \n"
            f"**Commit:** `{git_info[1]}`   \n"
            f"**Repo:** {git_info[3]}  \n"
            f"**Status:** {git_info[2]} \n"
        )
    out_lines.append("## Index\n")
    index_tree = group_files_by_folder(files_list)
    out_lines.extend(render_index(index_tree, "", git_info, level=0))
    out_lines.append("\n---\n")
    ordered_files = flatten_files_in_tree(index_tree)
    for file_path in ordered_files:
        ext = os.path.splitext(file_path)[1].lower()
        language = LANG_EXT.get(ext, "")
        anchor = file_path.replace("/", "").replace(".", "").replace("\\", "")
        out_lines.append(f"## {file_path}\n<a name=\"{anchor}\"></a>\n")
        abs_path = file_path if not base_docs_path else os.path.join(base_docs_path, file_path)
        if os.path.exists(abs_path) and not is_binary(abs_path):
            try:
                with open(abs_path, 'r', encoding='utf-8', errors='replace') as f:
                    total_lines = sum(1 for _ in f)
                out_lines.append(f"_(Total lines: {total_lines})_\n")
            except Exception:
                pass
        if is_binary(abs_path):
            out_lines.append("_(Binary file omitted)_\n")
            out_lines.append(f"```{language}\n```\n")
        elif ext == ".csv":
            out_lines.append("_(CSV preview)_\n")
            out_lines.append(f"_(Showing first {config['max_log_lines']} rows)_\n")
            out_lines.extend(preview_csv(abs_path, config))
            out_lines.append("\n")
        elif ext == ".xlsx":
            out_lines.append("_(Excel preview)_\n")
            out_lines.append(f"_(Showing first {config['max_log_lines']} rows)_\n")
            out_lines.extend(preview_excel(abs_path, config))
            out_lines.append("\n")
        elif ext == ".xls":
            out_lines.append("_(Preview not supported for .xls files. Use .xlsx)_\n")    
        elif ext == ".log":
            try:
                with open(abs_path, 'r', encoding='utf-8', errors='replace') as fin:
                    lines = fin.readlines()
                if len(lines) > config["max_log_lines"]:
                    out_lines.append(f"_(Showing last {config['max_log_lines']} lines)_\n")
                numbered_lines = add_line_numbers(lines[-config["max_log_lines"]:], start=max(1, len(lines)-config["max_log_lines"]+1))
                out_lines.append(f"```{language}\n")
                out_lines.extend(numbered_lines)
                out_lines.append("```\n")
            except Exception as e: 
                logging.warning(f"Could not read {abs_path}: {e}")
                out_lines.append(f"(Could not read file: {e})\n")
        elif should_truncate_file(abs_path, config):
            try:
                lines = read_n_lines_max_chars(abs_path, config)
                if not no_lines:
                    lines = add_line_numbers(lines)
                else:
                    lines = [l if l.endswith('\n') else l + '\n' for l in lines]
                out_lines.append(f"_(Showing up to {config['truncate_lines']} lines, max {config['max_truncate_chars']} characters)_\n")
                out_lines.append(f"```{language}\n")
                out_lines.extend(lines)
                out_lines.append("```\n")
            except Exception as e:
                logging.warning(f"Could not read {abs_path}: {e}")
                out_lines.append(f"(Could not read file: {e})\n")
        else:
            try:
                with open(abs_path, 'r', encoding='utf-8', errors='replace') as fin:
                    lines = fin.readlines()
                if not no_lines:
                    lines = add_line_numbers(lines)
                else:
                    lines = [l if l.endswith('\n') else l + '\n' for l in lines]                    
                out_lines.append(f"```{language}\n")
                out_lines.extend(lines)
                out_lines.append("```\n")
            except Exception as e:
                logging.warning(f"Could not read {abs_path}: {e}")
                out_lines.append(f"(Could not read file: {e})\n")        
        if not out_lines[-1].endswith('\n'):
            out_lines.append('\n')        
        out_lines.append("---\n")
    with open(output_file, 'w', encoding='utf-8') as out:
        out.write("".join(out_lines))


def main():
    parser = ArgumentParser(description="Generates directory tree and file content documentation in Markdown, with Git info, links, and special handling for binaries, logs, and truncated files.")
    parser.add_argument("directory", help="Project root directory.")
    parser.add_argument("--config", default="config.json", help="JSON configuration file.")
    parser.add_argument("--max-depth", type=int, help="Maximum tree depth.")
    parser.add_argument("--tree-title", default="Project Directory Tree", help="Title for the tree Markdown.")
    parser.add_argument("--truncate-lines", type=int, help="Max lines for files like README, LICENSE, etc.")
    parser.add_argument("--truncate-chars", type=int, help="Max characters to show for truncated files")
    parser.add_argument("--max-log-lines", type=int, help="Max lines to show for .log files")
    parser.add_argument("--max-preview-columns", type=int, help="Max columns to show in CSV/Excel preview")
    parser.add_argument("--no-lines", action="store_true", help="Disable per-line numbering in file previews")
    args = parser.parse_args()
    
    config = load_config_with_defaults(args.config, args)

    ensure_gitignore_has_internal_docs(args.directory)
    output_dir = ensure_internal_docs_dir(args.directory)
    tree_output_path = os.path.join(output_dir, "directory_tree.md")
    content_output_path = os.path.join(output_dir, "all_files_content.md")
    files_list, git_info = print_tree_and_collect_files(
        args.directory,
        config,
        output_file=tree_output_path,
        tree_title=args.tree_title
    )
    generate_content_document(
        files_list,
        output_file=content_output_path,
        git_info=git_info,
        config=config,
        base_docs_path=args.directory,
        project_dir=args.directory,
        no_lines=args.no_lines
    )
    print(f"DONE -> Directory tree saved in {tree_output_path}")
    print(f"DONE -> File contents saved in {content_output_path}")